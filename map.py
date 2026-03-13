import pandas as pd
from datetime import datetime
import pdfplumber
import re
import sys
import os
import glob

def clean_and_convert_to_float(value):
    """
    Cleans a string to extract a number and converts it to a float.
    This is a robust function that handles extra characters, symbols, and 
    non-numeric text by stripping them out before conversion.
    """
    if value is None:
        return None
    
    text = str(value)
    
    # Remove any character that is not a digit, a decimal point, or a minus sign.
    cleaned_text = re.sub(r'[^\d.-]', '', text)
    
    # If the cleaned string is empty or just a symbol, it's invalid.
    if not cleaned_text or cleaned_text in ['.', '-']:
        return None
        
    try:
        # Convert the cleaned string to a float.
        return float(cleaned_text)
    except (ValueError, TypeError):
        # If conversion fails for any reason, return None.
        return None

def extract_value_from_combined_cell(cell_text):
    """
    Extract name and numeric value from a combined cell like 'Brazil 3.75'
    Returns: (name, value) tuple
    """
    if not cell_text:
        return None, None

    text = str(cell_text).strip()
    # Try to split on last space to separate name from number
    parts = text.rsplit(' ', 1)
    if len(parts) == 2:
        name = parts[0].replace('\n', ' ').strip()
        value = clean_and_convert_to_float(parts[1])
        return name, value
    return text.replace('\n', ' ').strip(), None

def identify_table_type(table):
    """
    Identify what type of table this is based on its content.
    Returns: 'portfolio_stats', 'country_duration', 'fx_weights', or None
    """
    if not table or len(table) == 0:
        return None

    # Convert first few rows to text for analysis
    header_text = ""
    for row in table[:3]:  # Check first 3 rows
        if row:
            header_text += " ".join([str(cell) for cell in row if cell])

    header_text = header_text.lower()

    # Portfolio stats: contains "yield to maturity"
    if "yield to maturity" in header_text or "yield" in header_text and "modified duration" in header_text:
        return 'portfolio_stats'

    # Country duration: contains "country" and "duration"
    if "country" in header_text and "duration" in header_text:
        return 'country_duration'

    # FX weights: contains "currency" and "fund"
    if "currency" in header_text and "fund" in header_text:
        return 'fx_weights'

    return None

def parse_data_from_document(doc_path):
    """
    Extracts data directly from a PDF document by finding and parsing tables.
    This is a ROBUST implementation that handles:
    - Tables in any order
    - Different column structures
    - Dynamic country/currency names

    Args:
        doc_path (str): The file path to the PDF document to be parsed.

    Returns:
        tuple: A tuple containing the date string, and dictionaries for
               portfolio stats, country duration, and FX weights.
    """
    print(f"INFO: Starting data extraction from '{doc_path}'...")

    date_str = None
    portfolio_stats = {}
    country_duration = {}
    fx_weights = {}

    try:
        with pdfplumber.open(doc_path) as pdf:
            full_text = ""
            for page in pdf.pages:
                full_text += page.extract_text(x_tolerance=1, y_tolerance=3)

            # --- Date Extraction ---
            date_match = re.search(r'(\d{1,2}\s\w{3}\s\d{4})', full_text)
            if date_match:
                date_str = date_match.group(1)
                print(f"INFO: Found date: {date_str}")
            else:
                print("WARN: Date not found in document.")

            # --- Table Extraction ---
            for page_num, page in enumerate(pdf.pages):
                tables = page.extract_tables(table_settings={"vertical_strategy": "lines", "horizontal_strategy": "lines"})
                for table in tables:
                    if not table: continue

                    # Identify table type dynamically
                    table_type = identify_table_type(table)

                    if not table_type:
                        continue

                    # === PORTFOLIO STATS TABLE ===
                    if table_type == 'portfolio_stats':
                        print(f"INFO: Found 'Portfolio stats' table on page {page_num + 1}")
                        for row in table:
                            if row and len(row) >= 2 and row[0] and row[1]:
                                # Clean the key: remove percentages, superscripts, trailing numbers
                                key = str(row[0]).replace(' (%)', '').replace('¹', '').replace(' 1', '').strip()
                                # Skip header rows
                                if 'as at' in key.lower() or not key:
                                    continue
                                value = clean_and_convert_to_float(row[1])
                                if value is not None:
                                    portfolio_stats[key] = value

                    # === COUNTRY DURATION TABLE ===
                    elif table_type == 'country_duration':
                        print(f"INFO: Found 'countries by duration' table on page {page_num + 1}")
                        for row in table[1:]:  # Skip header row
                            if len(row) >= 2 and row[0]:
                                # Extract country and duration from combined or separate cells
                                country, duration = extract_value_from_combined_cell(row[0])

                                # Get benchmark value
                                benchmark = clean_and_convert_to_float(row[1]) if len(row) >= 2 else None

                                # Store data if valid
                                if country and duration is not None and benchmark is not None:
                                    country_duration[country] = [duration, benchmark]

                    # === FX WEIGHTS TABLE ===
                    elif table_type == 'fx_weights':
                        print(f"INFO: Found 'FX weights' table on page {page_num + 1}")
                        for row in table[1:]:  # Skip header row
                            if len(row) >= 2 and row[0]:
                                # Extract currency and fund % from combined or separate cells
                                currency, fund = extract_value_from_combined_cell(row[0])

                                # Get benchmark value
                                benchmark = clean_and_convert_to_float(row[1]) if len(row) >= 2 else None

                                # Store data if valid
                                if currency and fund is not None and benchmark is not None:
                                    fx_weights[currency] = [fund, benchmark]
                                    
    except FileNotFoundError:
        print(f"ERROR: The file was not found at the specified path: {doc_path}")
        return None, {}, {}, {}
    except Exception as e:
        print(f"An unexpected error occurred: {e}")
        return None, {}, {}, {}

    return date_str, portfolio_stats, country_duration, fx_weights

def process_and_map_data(date_str, portfolio_stats, country_duration, fx_weights):
    """
    Takes the extracted raw data, processes it, maps it to the data codes,
    and uses a hardcoded descriptive header for the final CSV output.
    """
    if not date_str:
        print("ERROR: Cannot process data without a date.")
        return pd.DataFrame(), [], []

    # =========================================================================
    # COLUMN REGISTRY — single source of truth for codes + descriptions.
    #
    # HOW TO ADD A NEW MEASURE:
    #   1. Pick the correct section (PORTSTST / COUNTRY / FX).
    #   2. Add ONE tuple: ('RMP.AVIVA.<SECTION>.<CODE>.M', '<Description>').
    #   3. For a new COUNTRY: also add its name→code entry to COUNTRY_MAP below.
    #   4. For a new FX currency: also add its name→code entry to CURRENCY_MAP below.
    #
    # Column code pattern:
    #   Portfolio stats  : RMP.AVIVA.PORTSTST.<FIELD>.M
    #   Country duration : RMP.AVIVA.COUNTRY.DUR.<ISO3>.M
    #   Country benchmark: RMP.AVIVA.COUNTRY.BENCH.<ISO3>.M
    #   FX fund weight   : RMP.AVIVA.FX.FUND.<CISO>.M   (currency ISO, e.g. TRY not TUR)
    #   FX benchmark     : RMP.AVIVA.FX.BENCH.<CISO>.M
    # =========================================================================

    _D  = 'Overweight And underweight countries by duration: Duration: '
    _B  = 'Overweight And underweight countries by duration: Relative to benchmark: '
    _FF = 'Overweights & underweights by FX: Fund: '
    _FB = 'Overweights & underweights by FX: Relative to benchmark: '

    COLUMN_REGISTRY = [
        # ── Portfolio stats ──────────────────────────────────────────────────
        ('RMP.AVIVA.PORTSTST.YIELD.M',     'Portfolio stats: Yield to maturity'),
        ('RMP.AVIVA.PORTSTST.MODDUR.M',    'Portfolio stats: Modified duration'),
        ('RMP.AVIVA.PORTSTST.TIMEMAT.M',   'Portfolio stats: Time to maturity'),
        ('RMP.AVIVA.PORTSTST.SPREADDUR.M', 'Portfolio stats: Spread duration'),

        # ── Country duration (fund) ──────────────────────────────────────────
        # To add a new country: add a row here AND add to COUNTRY_MAP below.
        ('RMP.AVIVA.COUNTRY.DUR.CZE.M',  _D + 'Czech Republic'),
        ('RMP.AVIVA.COUNTRY.DUR.USA.M',  _D + 'United States'),
        ('RMP.AVIVA.COUNTRY.DUR.POL.M',  _D + 'Poland'),
        ('RMP.AVIVA.COUNTRY.DUR.BRA.M',  _D + 'Brazil'),
        ('RMP.AVIVA.COUNTRY.DUR.UKR.M',  _D + 'Ukraine'),
        ('RMP.AVIVA.COUNTRY.DUR.IND.M',  _D + 'India'),
        ('RMP.AVIVA.COUNTRY.DUR.IDN.M',  _D + 'Indonesia'),
        ('RMP.AVIVA.COUNTRY.DUR.CHL.M',  _D + 'Chile'),
        ('RMP.AVIVA.COUNTRY.DUR.MYS.M',  _D + 'Malaysia'),
        ('RMP.AVIVA.COUNTRY.DUR.EURP.M', _D + 'European Union'),
        ('RMP.AVIVA.COUNTRY.DUR.THA.M',  _D + 'Thailand'),
        ('RMP.AVIVA.COUNTRY.DUR.TUR.M',  _D + 'Turkey'),
        ('RMP.AVIVA.COUNTRY.DUR.ECU.M',  _D + 'Ecuador'),
        ('RMP.AVIVA.COUNTRY.DUR.EGY.M',  _D + 'Egypt'),
        ('RMP.AVIVA.COUNTRY.DUR.PER.M',  _D + 'Peru'),
        ('RMP.AVIVA.COUNTRY.DUR.MEX.M',  _D + 'Mexico'),
        ('RMP.AVIVA.COUNTRY.DUR.DOM.M',  _D + 'Dominican Republic'),
        ('RMP.AVIVA.COUNTRY.DUR.CHN.M',  _D + 'China'),
        ('RMP.AVIVA.COUNTRY.DUR.ZAF.M',  _D + 'South Africa'),
        ('RMP.AVIVA.COUNTRY.DUR.COL.M',  _D + 'Colombia'),
        ('RMP.AVIVA.COUNTRY.DUR.ROU.M',  _D + 'Romania'),
        ('RMP.AVIVA.COUNTRY.DUR.URY.M',  _D + 'Uruguay'),

        # ── Country duration (relative to benchmark) ─────────────────────────
        ('RMP.AVIVA.COUNTRY.BENCH.CZE.M',  _B + 'Czech Republic'),
        ('RMP.AVIVA.COUNTRY.BENCH.USA.M',  _B + 'United States'),
        ('RMP.AVIVA.COUNTRY.BENCH.POL.M',  _B + 'Poland'),
        ('RMP.AVIVA.COUNTRY.BENCH.BRA.M',  _B + 'Brazil'),
        ('RMP.AVIVA.COUNTRY.BENCH.UKR.M',  _B + 'Ukraine'),
        ('RMP.AVIVA.COUNTRY.BENCH.IND.M',  _B + 'India'),
        ('RMP.AVIVA.COUNTRY.BENCH.IDN.M',  _B + 'Indonesia'),
        ('RMP.AVIVA.COUNTRY.BENCH.CHL.M',  _B + 'Chile'),
        ('RMP.AVIVA.COUNTRY.BENCH.MYS.M',  _B + 'Malaysia'),
        ('RMP.AVIVA.COUNTRY.BENCH.EURP.M', _B + 'European Union'),
        ('RMP.AVIVA.COUNTRY.BENCH.THA.M',  _B + 'Thailand'),
        ('RMP.AVIVA.COUNTRY.BENCH.TUR.M',  _B + 'Turkey'),
        ('RMP.AVIVA.COUNTRY.BENCH.ECU.M',  _B + 'Ecuador'),
        ('RMP.AVIVA.COUNTRY.BENCH.EGY.M',  _B + 'Egypt'),
        ('RMP.AVIVA.COUNTRY.BENCH.PER.M',  _B + 'Peru'),
        ('RMP.AVIVA.COUNTRY.BENCH.MEX.M',  _B + 'Mexico'),
        ('RMP.AVIVA.COUNTRY.BENCH.DOM.M',  _B + 'Dominican Republic'),
        ('RMP.AVIVA.COUNTRY.BENCH.CHN.M',  _B + 'China'),
        ('RMP.AVIVA.COUNTRY.BENCH.ZAF.M',  _B + 'South Africa'),
        ('RMP.AVIVA.COUNTRY.BENCH.COL.M',  _B + 'Colombia'),
        ('RMP.AVIVA.COUNTRY.BENCH.ROU.M',  _B + 'Romania'),
        ('RMP.AVIVA.COUNTRY.BENCH.URY.M',  _B + 'Uruguay'),

        # ── FX weights (fund) — NOTE: code uses ISO currency ticker, not country ──
        # To add a new currency: add a row here AND add to CURRENCY_MAP below.
        ('RMP.AVIVA.FX.FUND.TRY.M', _FF + 'Turkish Lira'),
        ('RMP.AVIVA.FX.FUND.USD.M', _FF + 'US Dollar'),
        ('RMP.AVIVA.FX.FUND.EGP.M', _FF + 'Egyptian Pound'),
        ('RMP.AVIVA.FX.FUND.COP.M', _FF + 'Colombian Peso'),
        ('RMP.AVIVA.FX.FUND.UYU.M', _FF + 'Uruguayan Peso'),
        ('RMP.AVIVA.FX.FUND.CZK.M', _FF + 'Czech Republic Koruna'),
        ('RMP.AVIVA.FX.FUND.THB.M', _FF + 'Thai Baht'),
        ('RMP.AVIVA.FX.FUND.CNY.M', _FF + 'Chinese Yuan'),
        ('RMP.AVIVA.FX.FUND.PLN.M', _FF + 'Polish Zloty'),
        ('RMP.AVIVA.FX.FUND.MYR.M', _FF + 'Malaysian Ringgit'),
        ('RMP.AVIVA.FX.FUND.INR.M', _FF + 'Indian Rupee'),
        ('RMP.AVIVA.FX.FUND.RON.M', _FF + 'Romanian Leu'),
        ('RMP.AVIVA.FX.FUND.BRL.M', _FF + 'Brazilian Real'),
        ('RMP.AVIVA.FX.FUND.ZAR.M', _FF + 'South African Rand'),
        ('RMP.AVIVA.FX.FUND.MXN.M', _FF + 'Mexican Peso'),
        ('RMP.AVIVA.FX.FUND.IDR.M', _FF + 'Indonesian Rupiah'),

        # ── FX weights (relative to benchmark) ──────────────────────────────
        ('RMP.AVIVA.FX.BENCH.TRY.M', _FB + 'Turkish Lira'),
        ('RMP.AVIVA.FX.BENCH.USD.M', _FB + 'US Dollar'),
        ('RMP.AVIVA.FX.BENCH.EGP.M', _FB + 'Egyptian Pound'),
        ('RMP.AVIVA.FX.BENCH.COP.M', _FB + 'Colombian Peso'),
        ('RMP.AVIVA.FX.BENCH.UYU.M', _FB + 'Uruguayan Peso'),
        ('RMP.AVIVA.FX.BENCH.CZK.M', _FB + 'Czech Republic Koruna'),
        ('RMP.AVIVA.FX.BENCH.THB.M', _FB + 'Thai Baht'),
        ('RMP.AVIVA.FX.BENCH.CNY.M', _FB + 'Chinese Yuan'),
        ('RMP.AVIVA.FX.BENCH.PLN.M', _FB + 'Polish Zloty'),
        ('RMP.AVIVA.FX.BENCH.MYR.M', _FB + 'Malaysian Ringgit'),
        ('RMP.AVIVA.FX.BENCH.INR.M', _FB + 'Indian Rupee'),
        ('RMP.AVIVA.FX.BENCH.RON.M', _FB + 'Romanian Leu'),
        ('RMP.AVIVA.FX.BENCH.BRL.M', _FB + 'Brazilian Real'),
        ('RMP.AVIVA.FX.BENCH.ZAR.M', _FB + 'South African Rand'),
        ('RMP.AVIVA.FX.BENCH.MXN.M', _FB + 'Mexican Peso'),
        ('RMP.AVIVA.FX.BENCH.IDR.M', _FB + 'Indonesian Rupiah'),
    ]

    all_data_codes      = [code for code, _   in COLUMN_REGISTRY]
    descriptive_headers = [''] + [desc for _, desc in COLUMN_REGISTRY]

    # =========================================================================
    # LOOKUP MAPS — full name → code used in column registry above.
    #
    # COUNTRY_MAP: country name (as it appears in PDF) → 3-letter country code
    # CURRENCY_MAP: currency name (as it appears in PDF) → ISO currency ticker
    # =========================================================================

    COUNTRY_MAP = {
        'Czech Republic': 'CZE', 'United States': 'USA', 'Poland': 'POL', 'Brazil': 'BRA',
        'Ukraine': 'UKR', 'India': 'IND', 'Indonesia': 'IDN', 'Chile': 'CHL',
        'Malaysia': 'MYS', 'European Union': 'EURP', 'South Africa': 'ZAF', 'Mexico': 'MEX',
        'Thailand': 'THA', 'Turkey': 'TUR', 'Ecuador': 'ECU', 'Egypt': 'EGY',
        'Peru': 'PER', 'Dominican Republic': 'DOM', 'China': 'CHN', 'Colombia': 'COL',
        'Romania': 'ROU', 'Uruguay': 'URY',
        # Additional countries (not yet in factsheet but mapped for future use):
        'Philippines': 'PHL', 'Russia': 'RUS', 'Hungary': 'HUN',
        'Nigeria': 'NGA', 'Kenya': 'KEN', 'Ghana': 'GHA', 'Morocco': 'MAR',
        'Argentina': 'ARG', 'Paraguay': 'PRY', 'Vietnam': 'VNM',
        'Pakistan': 'PAK', 'Bangladesh': 'BGD', 'Sri Lanka': 'LKA',
    }

    CURRENCY_MAP = {
        # Name as it appears in PDF → ISO currency ticker
        'Turkish Lira': 'TRY', 'US Dollar': 'USD', 'Egyptian Pound': 'EGP',
        'Colombian Peso': 'COP', 'Uruguayan Peso': 'UYU', 'Czech Republic Koruna': 'CZK',
        'Thai Baht': 'THB', 'Chinese Yuan': 'CNY', 'Polish Zloty': 'PLN',
        'Malaysian Ringgit': 'MYR', 'Indian Rupee': 'INR', 'Romanian Leu': 'RON',
        'Brazilian Real': 'BRL', 'South African Rand': 'ZAR', 'Mexican Peso': 'MXN',
        'Indonesian Rupiah': 'IDR',
        # Additional currencies (not yet in factsheet but mapped for future use):
        'Philippine Peso': 'PHP', 'Russian Ruble': 'RUB', 'Hungarian Forint': 'HUF',
        'Nigerian Naira': 'NGN', 'Kenyan Shilling': 'KES', 'Argentine Peso': 'ARS',
        'Chilean Peso': 'CLP', 'Vietnamese Dong': 'VND', 'Pakistani Rupee': 'PKR',
        'Bangladeshi Taka': 'BDT', 'Sri Lankan Rupee': 'LKR',
    }

    def generate_code_from_name(name):
        """Fallback: derive a code from name if not in map. Log a warning so it can be added."""
        return name.replace(' ', '')[:3].upper()

    dt_object = datetime.strptime(date_str, "%d %b %Y")
    time_period = dt_object.strftime("%Y-%m")

    available_data = {
        'RMP.AVIVA.PORTSTST.YIELD.M':     portfolio_stats.get('Yield to maturity'),
        'RMP.AVIVA.PORTSTST.MODDUR.M':    portfolio_stats.get('Modified duration'),
        'RMP.AVIVA.PORTSTST.TIMEMAT.M':   portfolio_stats.get('Time to maturity'),
        'RMP.AVIVA.PORTSTST.SPREADDUR.M': portfolio_stats.get('Spread duration'),
    }

    # Tracks any names found in the PDF that are not yet in the registry.
    # These are still included in the output data automatically.
    new_measures = []   # list of dicts: {type, name, code, dur_code, bench_code, description}

    for country, values in country_duration.items():
        in_map = country in COUNTRY_MAP
        code = COUNTRY_MAP.get(country, generate_code_from_name(country))
        dur_col   = f'RMP.AVIVA.COUNTRY.DUR.{code}.M'
        bench_col = f'RMP.AVIVA.COUNTRY.BENCH.{code}.M'
        available_data[dur_col]   = values[0]
        available_data[bench_col] = values[1]
        if dur_col not in all_data_codes:
            msg = f"New country detected (included in output): {country} -> {code}"
            print(f"INFO: {msg}")
            new_measures.append({
                'Type': 'Country',
                'Name in PDF': country,
                'Code': code,
                'Suggested COUNTRY_MAP entry': f"'{country}': '{code}'",
                'Suggested COLUMN_REGISTRY entries (DUR)':   f"('RMP.AVIVA.COUNTRY.DUR.{code}.M',  _D + '{country}')",
                'Suggested COLUMN_REGISTRY entries (BENCH)': f"('RMP.AVIVA.COUNTRY.BENCH.{code}.M', _B + '{country}')",
                'Values found': f"duration={values[0]}, benchmark={values[1]}",
            })
            # Auto-extend the column list so the data is written
            if not in_map:
                all_data_codes.append(dur_col)
                all_data_codes.append(bench_col)

    for currency, values in fx_weights.items():
        in_map = currency in CURRENCY_MAP
        code = CURRENCY_MAP.get(currency, generate_code_from_name(currency))
        fund_col  = f'RMP.AVIVA.FX.FUND.{code}.M'
        bench_col = f'RMP.AVIVA.FX.BENCH.{code}.M'
        available_data[fund_col]  = values[0]
        available_data[bench_col] = values[1]
        if fund_col not in all_data_codes:
            msg = f"New currency detected (included in output): {currency} -> {code}"
            print(f"INFO: {msg}")
            new_measures.append({
                'Type': 'Currency (FX)',
                'Name in PDF': currency,
                'Code': code,
                'Suggested CURRENCY_MAP entry': f"'{currency}': '{code}'",
                'Suggested COLUMN_REGISTRY entries (FUND)':  f"('RMP.AVIVA.FX.FUND.{code}.M',  _FF + '{currency}')",
                'Suggested COLUMN_REGISTRY entries (BENCH)': f"('RMP.AVIVA.FX.BENCH.{code}.M', _FB + '{currency}')",
                'Values found': f"fund={values[0]}, benchmark={values[1]}",
            })
            if not in_map:
                all_data_codes.append(fund_col)
                all_data_codes.append(bench_col)

    final_data_row = {'Time Period': time_period}
    for code in all_data_codes:
        final_data_row[code] = available_data.get(code)
    df = pd.DataFrame([final_data_row])

    return df, descriptive_headers, all_data_codes, new_measures

DOWNLOAD_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'downloads')
OUTPUT_DIR   = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'output')
os.makedirs(OUTPUT_DIR, exist_ok=True)

def find_pdf_files(directory=DOWNLOAD_DIR):
    """
    Find all PDF files in the given directory.
    Returns a list of PDF file paths.
    """
    pdf_files = glob.glob(os.path.join(directory, '*.pdf'))
    return pdf_files

def write_excel_output(combined_df, descriptive_headers, data_codes, all_new_measures, output_path):
    """
    Write the extracted data to an Excel file with two sheets:
      - 'Data'         : two-header rows (codes + descriptions) then data rows
      - 'New Measures' : any countries/currencies auto-detected but not in the registry
    """
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:

        # ── Sheet 1: Data ────────────────────────────────────────────────────
        # Build a display DataFrame with descriptive headers as row 1
        code_header_row = pd.DataFrame(
            [['Time Period'] + data_codes],
            columns=combined_df.columns
        )
        desc_header_row = pd.DataFrame(
            [descriptive_headers],
            columns=combined_df.columns
        )
        display_df = pd.concat(
            [code_header_row, desc_header_row, combined_df],
            ignore_index=True
        )
        display_df.to_excel(writer, sheet_name='Data', index=False, header=False)

        ws = writer.sheets['Data']
        # Bold the first two header rows
        from openpyxl.styles import Font, PatternFill, Alignment
        header_font  = Font(bold=True)
        header_fill  = PatternFill(fill_type='solid', fgColor='D9E1F2')
        for row in ws.iter_rows(min_row=1, max_row=2):
            for cell in row:
                cell.font  = header_font
                cell.fill  = header_fill
                cell.alignment = Alignment(wrap_text=True)
        ws.row_dimensions[1].height = 18
        ws.row_dimensions[2].height = 40

        # ── Sheet 2: New Measures ────────────────────────────────────────────
        if all_new_measures:
            new_df = pd.DataFrame(all_new_measures)
            new_df.to_excel(writer, sheet_name='New Measures', index=False)

            ws2 = writer.sheets['New Measures']
            warn_font = Font(bold=True, color='C00000')
            warn_fill = PatternFill(fill_type='solid', fgColor='FCE4D6')
            for cell in ws2[1]:
                cell.font  = warn_font
                cell.fill  = warn_fill
            # Auto-fit column widths
            for col in ws2.columns:
                max_len = max(len(str(c.value or '')) for c in col)
                ws2.column_dimensions[col[0].column_letter].width = min(max_len + 4, 80)

            print(f"\n{'!'*70}")
            print(f"  {len(all_new_measures)} NEW MEASURE(S) DETECTED — see 'New Measures' sheet in Excel")
            print(f"  To permanently register them, add the entries shown in that sheet")
            print(f"  to COLUMN_REGISTRY and COUNTRY_MAP / CURRENCY_MAP in map.py")
            print(f"{'!'*70}")
        else:
            # Write an empty sheet with a note
            pd.DataFrame([{'Status': 'No new measures detected — all data matched the registry.'}]) \
              .to_excel(writer, sheet_name='New Measures', index=False)

    print(f"Excel output saved to: {output_path}")


def process_all_pdfs(directory=DOWNLOAD_DIR):
    """
    Process all PDF files in the directory and write an Excel file with two sheets:
    'Data' (main results) and 'New Measures' (auto-detected entries not in registry).
    """
    pdf_files = find_pdf_files(directory)

    if not pdf_files:
        print("ERROR: No PDF files found in directory")
        return

    print(f"INFO: Found {len(pdf_files)} PDF file(s) to process")

    all_data_rows    = []
    all_new_measures = []
    descriptive_headers = None
    data_codes = None

    for pdf_file in pdf_files:
        print(f"\n{'='*70}")
        print(f"Processing: {os.path.basename(pdf_file)}")
        print(f"{'='*70}")

        date, p_stats, c_dur, fx_w = parse_data_from_document(pdf_file)

        if date:
            extracted_df, desc_headers, codes, new_measures = process_and_map_data(
                date, p_stats, c_dur, fx_w
            )
            if descriptive_headers is None:
                descriptive_headers = desc_headers
                data_codes = codes

            all_data_rows.append(extracted_df)
            all_new_measures.extend(new_measures)
            print(f"SUCCESS: Extracted data for {date}")
        else:
            print(f"FAILED: Could not extract data from {os.path.basename(pdf_file)}")

    if all_data_rows:
        combined_df = pd.concat(all_data_rows, ignore_index=True).sort_values('Time Period')
        output_path = os.path.join(OUTPUT_DIR, "RMP_AVIVA_DATA_EXTRACTED.xlsx")
        write_excel_output(combined_df, descriptive_headers, data_codes, all_new_measures, output_path)

        print("\n" + "="*70)
        print("--- Script Finished ---")
        print(f"Processed {len(all_data_rows)} PDF file(s) | Total rows: {len(combined_df)}")
        print("="*70)
    else:
        print("\n--- Script Aborted ---")
        print("Could not extract data from any PDF files.")


if __name__ == "__main__":
    if len(sys.argv) > 1:
        arg = sys.argv[1]

        if arg.endswith('.pdf') and os.path.isfile(arg):
            print(f"Processing single file: {arg}")
            date, p_stats, c_dur, fx_w = parse_data_from_document(arg)

            if date:
                extracted_df, descriptive_headers, data_codes, new_measures = process_and_map_data(
                    date, p_stats, c_dur, fx_w
                )
                output_path = os.path.join(OUTPUT_DIR, "RMP_AVIVA_DATA_EXTRACTED.xlsx")
                write_excel_output(extracted_df, descriptive_headers, data_codes, new_measures, output_path)
                print("\n--- Script Finished ---")
            else:
                print("\n--- Script Aborted ---")
                print("Could not extract the necessary data to proceed.")

        elif os.path.isdir(arg):
            print(f"Processing all PDFs in directory: {arg}")
            process_all_pdfs(arg)
        else:
            print(f"ERROR: '{arg}' is not a valid file or directory")
    else:
        print(f"Processing all PDFs in {DOWNLOAD_DIR}...")
        process_all_pdfs()

